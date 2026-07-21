"""
Data Manager Module for PJ Cost Record Application
Handles reading from Excel (03-1 PJ Cost Record update 202504_10.xlsx),
parsing data, normalizing JSON structures, persistent DB storage, and CRUD operations.
"""

import os
import json
import datetime
import openpyxl
from cost_engine import calculate_project_metrics

DB_PATH = os.path.join(os.path.dirname(__file__), "data", "db.json")
EXCEL_PATH = os.path.join(os.path.dirname(__file__), "03-1 PJ Cost Record update 202504_10.xlsx")

DEFAULT_CATEGORIES = [
    {"code": 10, "name": "MATERIALS"},
    {"code": 20, "name": "SUB CONTRACTOR"},
    {"code": 30, "name": "HEAVY EQUIPMENT"},
    {"code": 40, "name": "RENTAL"},
    {"code": 50, "name": "STAFF EXPENSE (Salary)"},
    {"code": 60, "name": "OFFICE EXPENSE"},
    {"code": 70, "name": "ENTERTAIN"},
    {"code": 80, "name": "MISCELLANEOUS"},
    {"code": 90, "name": "OTHER"}
]

def safe_float(val, default=0.0):
    if val is None:
        return default
    if isinstance(val, (int, float)):
        return float(val)
    try:
        # Replace commas if formatted string
        cleaned = str(val).replace(",", "").strip()
        return float(cleaned)
    except (ValueError, TypeError):
        return default

def format_date(val):
    if isinstance(val, (datetime.datetime, datetime.date)):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, str):
        return val
    return ""

def parse_excel_to_db():
    if not os.path.exists(EXCEL_PATH):
        print(f"Warning: {EXCEL_PATH} not found. Returning empty dataset.")
        return {"categories": DEFAULT_CATEGORIES, "projects": [], "forecasts": []}

    wb_val = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    wb_form = openpyxl.load_workbook(EXCEL_PATH, data_only=False)

    # 1. Parse Categories from 'JPP Code '
    categories = list(DEFAULT_CATEGORIES)
    if "JPP Code " in wb_val.sheetnames:
        ws_jpp = wb_val["JPP Code "]
        extracted_cat = []
        for r in range(4, ws_jpp.max_row + 1):
            c_code = ws_jpp.cell(r, 4).value or ws_jpp.cell(r, 2).value
            c_desc = ws_jpp.cell(r, 3).value
            if c_code and c_desc:
                try:
                    extracted_cat.append({"code": int(c_code), "name": str(c_desc).strip()})
                except ValueError:
                    pass
        if extracted_cat:
            categories = extracted_cat

    # 2. Parse Forecasts from 'Money forecast'
    forecasts = []
    if "Money forecast" in wb_val.sheetnames:
        ws_mf = wb_val["Money forecast"]
        for r in range(2, ws_mf.max_row + 1):
            seq = ws_mf.cell(r, 1).value
            title = ws_mf.cell(r, 2).value
            date_val = format_date(ws_mf.cell(r, 3).value)
            amt1 = safe_float(ws_mf.cell(r, 4).value)
            amt2 = safe_float(ws_mf.cell(r, 5).value)
            amt3 = safe_float(ws_mf.cell(r, 6).value)
            amt4 = safe_float(ws_mf.cell(r, 7).value)
            total = safe_float(ws_mf.cell(r, 8).value, amt1 + amt2 + amt3 + amt4)
            if title or amt1 or amt2 or amt3 or amt4:
                forecasts.append({
                    "id": r,
                    "title": str(title).strip() if title else f"Item {seq or r}",
                    "date": date_val,
                    "month_1": amt1,
                    "month_2": amt2,
                    "month_3": amt3,
                    "month_4": amt4,
                    "total": total
                })

    # 3. Parse Master Job Balance List from 'All Job balance report'
    master_jobs = {}
    if "All Job balance report" in wb_val.sheetnames:
        ws_bal = wb_val["All Job balance report"]
        for r in range(2, ws_bal.max_row + 1):
            job_code = ws_bal.cell(r, 2).value
            if job_code:
                job_code_str = str(job_code).strip()
                offer_no = str(ws_bal.cell(r, 3).value or "").strip()
                client = str(ws_bal.cell(r, 4).value or "").strip()
                pj_name = str(ws_bal.cell(r, 5).value or "").strip()
                desc = str(ws_bal.cell(r, 6).value or "").strip()
                cp = safe_float(ws_bal.cell(r, 7).value)
                bg = safe_float(ws_bal.cell(r, 8).value)

                master_jobs[job_code_str] = {
                    "job_code": job_code_str,
                    "offer_no": offer_no,
                    "client_name": client,
                    "project_name": pj_name,
                    "description": desc,
                    "contract_price": cp,
                    "budget_price": bg
                }

    # 4. Parse Individual Project Sheets
    projects = []
    project_sheets = [
        s for s in wb_val.sheetnames
        if s not in ["JPP Code ", "Money forecast", "Project Onhand", "All Job balance report"]
    ]

    for idx, sheet_name in enumerate(project_sheets):
        ws_v = wb_val[sheet_name]

        # Extract Header Metadata
        client_val = ws_v.cell(1, 3).value or ""
        pj_name_val = ws_v.cell(2, 3).value or ""
        job_code_val = ws_v.cell(3, 3).value or ""
        quotation_no = ws_v.cell(4, 3).value or ws_v.cell(4, 4).value or ""
        contract_price_val = ws_v.cell(5, 3).value or 0.0
        budget_price_val = ws_v.cell(6, 3).value or 0.0
        start_date = format_date(ws_v.cell(1, 9).value)

        job_code_clean = str(job_code_val).strip() if job_code_val and str(job_code_val).strip() != "=C30" else ""
        if not job_code_clean:
            # Fallback to sheet name code prefix
            job_code_clean = sheet_name.split()[0]

        # Merge with master_jobs if available
        master_info = master_jobs.get(job_code_clean, {})
        client_name = master_info.get("client_name") or str(client_val).strip() or "Client"
        project_name = master_info.get("project_name") or str(pj_name_val).strip() or sheet_name
        description = master_info.get("description") or ""
        offer_no = master_info.get("offer_no") or str(quotation_no).strip()
        contract_price = safe_float(master_info.get("contract_price") or contract_price_val)
        budget_price = safe_float(master_info.get("budget_price") or budget_price_val)

        # Extract Expense Items (Rows 10..24)
        expense_items = []
        for r in range(10, 25):
            vendor = ws_v.cell(r, 3).value
            mat_num = ws_v.cell(r, 4).value
            item_desc = ws_v.cell(r, 5).value
            text_quo = ws_v.cell(r, 6).value
            amount = ws_v.cell(r, 7).value
            updated_by = ws_v.cell(r, 8).value
            item_date = format_date(ws_v.cell(r, 9).value)
            po_val = ws_v.cell(r, 10).value if ws_v.max_column >= 10 else ""

            if vendor or mat_num or item_desc or (amount and amount != 0):
                try:
                    mat_code = int(mat_num) if mat_num is not None else 90
                except (ValueError, TypeError):
                    mat_code = 90

                expense_items.append({
                    "item_id": f"item_{r}",
                    "vendor": str(vendor).strip() if vendor else "",
                    "material_number": mat_code,
                    "description": str(item_desc).strip() if item_desc else "",
                    "text_quo": str(text_quo).strip() if text_quo else "",
                    "amount": safe_float(amount),
                    "updated_by": str(updated_by).strip() if updated_by else "",
                    "date": item_date,
                    "po": str(po_val).strip() if po_val else ""
                })

        # Extract Category Budget Breakdown (Rows 34..42 if exists)
        budget_categories = []
        for r in range(34, 43):
            cat_code = ws_v.cell(r, 3).value
            cat_desc = ws_v.cell(r, 5).value
            cat_bg = ws_v.cell(r, 6).value
            if cat_code and cat_desc:
                try:
                    code_int = int(cat_code)
                    budget_categories.append({
                        "code": code_int,
                        "description": str(cat_desc).strip(),
                        "budget": safe_float(cat_bg),
                        "expense": 0.0,
                        "balance": 0.0
                    })
                except (ValueError, TypeError):
                    pass

        if not budget_categories:
            # Provide default category structure
            for c in categories:
                budget_categories.append({
                    "code": c["code"],
                    "description": c["name"],
                    "budget": 0.0,
                    "expense": 0.0,
                    "balance": 0.0
                })

        project_obj = {
            "id": f"PJ-{idx+1:03d}",
            "job_code": job_code_clean,
            "offer_no": offer_no,
            "client_name": client_name,
            "project_name": project_name,
            "description": description,
            "sheet_name": sheet_name,
            "contract_price": contract_price,
            "budget_price": budget_price,
            "status": "In Progress" if (idx % 2 == 0) else "Completed",
            "start_date": start_date or "2025-10-01",
            "expense_items": expense_items,
            "budget_categories": budget_categories
        }

        # Calculate metrics using formula engine
        calculated_pj = calculate_project_metrics(project_obj)
        project_obj.update(calculated_pj)
        projects.append(project_obj)

    db_data = {
        "categories": categories,
        "projects": projects,
        "forecasts": forecasts,
        "last_updated": datetime.datetime.now().isoformat()
    }

    # Ensure data dir exists and save
    os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)
    with open(DB_PATH, "w", encoding="utf-8") as f:
        json.dump(db_data, f, ensure_ascii=False, indent=2)

    print(f"Loaded and saved {len(projects)} projects, {len(categories)} categories to {DB_PATH}")
    return db_data

def get_db():
    if not os.path.exists(DB_PATH):
        return parse_excel_to_db()
    try:
        with open(DB_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return parse_excel_to_db()

def save_db(db_data):
    db_data["last_updated"] = datetime.datetime.now().isoformat()
    os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)
    with open(DB_PATH, "w", encoding="utf-8") as f:
        json.dump(db_data, f, ensure_ascii=False, indent=2)

def get_project_by_id(project_id):
    db = get_db()
    for pj in db.get("projects", []):
        if pj["id"] == project_id:
            return pj
    return None

def save_project(project_data):
    db = get_db()
    projects = db.get("projects", [])

    # Calculate financial metrics before saving
    calc_data = calculate_project_metrics(project_data)
    project_data.update(calc_data)

    existing_idx = next((i for i, p in enumerate(projects) if p["id"] == project_data["id"]), None)
    if existing_idx is not None:
        projects[existing_idx] = project_data
    else:
        if not project_data.get("id"):
            project_data["id"] = f"PJ-{len(projects)+1:03d}"
        projects.append(project_data)

    db["projects"] = projects
    save_db(db)
    return project_data

def delete_project(project_id):
    db = get_db()
    projects = [p for p in db.get("projects", []) if p["id"] != project_id]
    db["projects"] = projects
    save_db(db)
    return True

if __name__ == "__main__":
    parse_excel_to_db()
